from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
from apps.projects.models import Project
from apps.provinces.models import District
from apps.sites.models import Site
from rest_framework.decorators import action
from rest_framework.response import Response
from apps.core.permissions import HasModulePermission
from .models import (
    DailyProgressEntry,
    ProgressTaskTemplate,
    SiteProgressTask,
    KPICategory,
    ModuleCatalogEntry,
    ItemCatalogEntry,
)
from .serializers import (
    DailyProgressEntrySerializer,
    ProgressTaskTemplateSerializer,
    SiteProgressTaskSerializer,
    KPICategorySerializer,
    ModuleCatalogEntrySerializer,
    ItemCatalogEntrySerializer,
)
import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.graphics.shapes import Drawing, String
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
from django.http import HttpResponse


class ProgressTaskTemplateViewSet(viewsets.ModelViewSet):
    """Admin catalog of reusable progress task / KPI types."""

    queryset = ProgressTaskTemplate.objects.all()
    serializer_class = ProgressTaskTemplateSerializer
    permission_classes = [IsAuthenticated, HasModulePermission]
    module_key = 'daily-progress'

    def get_queryset(self):
        qs = super().get_queryset()
        active = self.request.query_params.get('is_active')
        if active in ('true', '1'):
            qs = qs.filter(is_active=True)
        elif active in ('false', '0'):
            qs = qs.filter(is_active=False)
        category = self.request.query_params.get('category')
        if category:
            qs = qs.filter(category=category)
        return qs


class SiteProgressTaskViewSet(viewsets.ModelViewSet):
    """
    Tasks scoped to project + district (site optional). Filterable by any
    combination of project_id, district_id, site_id.
    """

    queryset = SiteProgressTask.objects.select_related(
        'project',
        'district',
        'site',
        'template',
        'boq',
        'boq_item',
    ).prefetch_related('entries')
    serializer_class = SiteProgressTaskSerializer
    permission_classes = [IsAuthenticated, HasModulePermission]
    module_key = 'daily-progress'

    def get_queryset(self):
        qs = super().get_queryset()
# In SiteProgressTaskViewSet.get_queryset, replace the loop with:
        for param in ('project_id', 'district_id', 'boq_id', 'boq_item_id'):
            value = self.request.query_params.get(param)
            if value:
                qs = qs.filter(**{param: value})

        site_id = self.request.query_params.get('site_id')
        if site_id == 'none':
            qs = qs.filter(site__isnull=True)
        elif site_id:
            qs = qs.filter(site_id=site_id)
        return qs

    def perform_create(self, serializer):
        instance = serializer.save()
        ItemCatalogEntry.objects.get_or_create(
            name=instance.name,
            defaults={'default_unit': instance.unit},
        )


class KPICategoryViewSet(viewsets.ModelViewSet):
    """
    Admin-defined Module groupings, scoped to project + district (site optional).
    """

    serializer_class = KPICategorySerializer
    permission_classes = [IsAuthenticated, HasModulePermission]
    module_key = 'daily-progress'

    def get_queryset(self):
        qs = KPICategory.objects.select_related('project', 'district', 'site').prefetch_related('subtasks')
        for param in ('project_id', 'district_id'):
            value = self.request.query_params.get(param)
            if value:
                qs = qs.filter(**{param: value})

        site_id = self.request.query_params.get('site_id')
        if site_id == 'none':
            qs = qs.filter(site__isnull=True)
        elif site_id:
            qs = qs.filter(site_id=site_id)
        return qs

    def perform_create(self, serializer):
        instance = serializer.save()
        ModuleCatalogEntry.objects.get_or_create(
            name=instance.name,
            defaults={'description': instance.description},
        )


class DailyProgressEntryViewSet(viewsets.ModelViewSet):
    queryset = DailyProgressEntry.objects.select_related(
        'site',
        'site_task',
        'site_task__project',
        'site_task__district',
        'site_task__boq',
        'site_task__boq_item',
        'submitted_by',
    )
    serializer_class = DailyProgressEntrySerializer
    permission_classes = [IsAuthenticated, HasModulePermission]
    module_key = 'daily-progress'

    def get_queryset(self):
        qs = super().get_queryset()
        for param in ('site_task_id', 'date'):
            value = self.request.query_params.get(param)
            if value:
                qs = qs.filter(**{param: value})
        site_id = self.request.query_params.get('site_id')
        if site_id == 'none':
            qs = qs.filter(site__isnull=True)
        elif site_id:
            qs = qs.filter(site_id=site_id)
        project_id = self.request.query_params.get('project_id')
        if project_id:
            qs = qs.filter(site_task__project_id=project_id)
        district_id = self.request.query_params.get('district_id')
        if district_id:
            qs = qs.filter(site_task__district_id=district_id)
        boq_id = self.request.query_params.get('boq_id')
        if boq_id:
            qs = qs.filter(site_task__boq_id=boq_id)
        boq_item_id = self.request.query_params.get('boq_item_id')
        if boq_item_id:
            qs = qs.filter(site_task__boq_item_id=boq_item_id)
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
        return qs

    def perform_create(self, serializer):
        entry = serializer.save(submitted_by=self.request.user)
        entry.site_task.sync_boq_actual()

    def perform_update(self, serializer):
        entry = serializer.save()
        entry.site_task.sync_boq_actual()

    def perform_destroy(self, instance):
        site_task = instance.site_task
        instance.delete()
        site_task.sync_boq_actual()


class ModuleCatalogViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = ModuleCatalogEntrySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = ModuleCatalogEntry.objects.all()
        q = self.request.query_params.get('q')
        if q:
            qs = qs.filter(name__icontains=q)
        return qs


class ItemCatalogViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = ItemCatalogEntrySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = ItemCatalogEntry.objects.all()
        q = self.request.query_params.get('q')
        if q:
            qs = qs.filter(name__icontains=q)
        return qs

class ProgressSummaryViewSet(viewsets.ViewSet):
    """
    Read-only rollups of progress at project / district / site level.
    Each level reports only its own scope — no blending across levels.
    """

    permission_classes = [IsAuthenticated, HasModulePermission]
    module_key = 'daily-progress'

    def _summarize(self, qs, name, extra=None):
        """
        Group tasks by activity (template, falling back to key) so units
        never get mixed (e.g. meters of fiber summed with camera units).
        Returns a headline weighted % plus a per-activity breakdown.
        """
        tasks = list(qs.select_related('template').prefetch_related('entries'))
        by_activity = {}
        for t in tasks:
            cumulative = sum((e.quantity for e in t.entries.all()), 0)
            planned = float(t.planned_quantity or 0)
            group_key = t.template_id or t.key
            if group_key not in by_activity:
                by_activity[group_key] = {
                    'label': t.name,
                    'unit': t.unit,
                    'planned': 0.0,
                    'actual': 0.0,
                }
            by_activity[group_key]['planned'] += planned
            by_activity[group_key]['actual'] += float(cumulative)

        breakdown = []
        for v in by_activity.values():
            pct = round((v['actual'] / v['planned']) * 100, 1) if v['planned'] else 0.0
            breakdown.append({**v, 'percent_complete': pct})
        breakdown.sort(key=lambda b: b['label'])

        total_planned = sum(v['planned'] for v in by_activity.values())
        total_actual = sum(v['actual'] for v in by_activity.values())
        overall_pct = round((total_actual / total_planned) * 100, 1) if total_planned else 0.0

        result = {
            'name': name,
            'tasks_count': len(tasks),
            'planned_total': total_planned,
            'actual_total': total_actual,
            'percent_complete': overall_pct,
            'breakdown': breakdown,
        }
        if extra:
            result.update(extra)
        return result

    @action(detail=False, methods=['get'], url_path='projects')
    def projects(self, request):
        """One row per project, its own scope only."""
        projects = Project.objects.all().order_by('name')
        data = [
            self._summarize(
                SiteProgressTask.objects.filter(project=p),
                p.name,
                extra={'project_id': str(p.id)},
            )
            for p in projects
        ]
        return Response({'data': data})

    @action(detail=False, methods=['get'], url_path=r'projects/(?P<project_id>[^/]+)/districts')
    def project_districts(self, request, project_id=None):
        """One row per district within a project — each district's own scope only."""
        district_ids = (
            SiteProgressTask.objects.filter(project_id=project_id)
            .values_list('district_id', flat=True)
            .distinct()
        )
        districts = District.objects.filter(id__in=district_ids).order_by('name')
        data = [
            self._summarize(
                SiteProgressTask.objects.filter(project_id=project_id, district=d),
                d.name,
                extra={'district_id': str(d.id)},
            )
            for d in districts
        ]
        return Response({'data': data, 'project_id': project_id})

    @action(detail=False, methods=['get'], url_path=r'districts/(?P<district_id>[^/]+)/sites')
    def district_sites(self, request, district_id=None):
        """
        One row per site within a district, PLUS a separate row for that
        district's own direct (no-site) tasks if any exist — never blended
        into the site rows.
        """
        site_ids = (
            SiteProgressTask.objects.filter(district_id=district_id, site__isnull=False)
            .values_list('site_id', flat=True)
            .distinct()
        )
        sites = Site.objects.filter(id__in=site_ids).order_by('name')
        data = [
            self._summarize(
                SiteProgressTask.objects.filter(district_id=district_id, site=s),
                s.name,
                extra={'site_id': str(s.id)},
            )
            for s in sites
        ]

        direct_qs = SiteProgressTask.objects.filter(district_id=district_id, site__isnull=True)
        if direct_qs.exists():
            direct_summary = self._summarize(
                direct_qs, 'District-wide (no specific site)', extra={'site_id': None}
            )
            data.append(direct_summary)

        return Response({'data': data, 'district_id': district_id})

    
    def _gather_export_rows(self, request):
        scope = request.query_params.get('scope', 'projects')
        include_breakdown = request.query_params.get('include_breakdown', 'true') == 'true'

        rows = []
        if scope == 'projects':
            project_ids = request.query_params.get('project_ids')
            projects = Project.objects.all().order_by('name')
            if project_ids:
                projects = projects.filter(id__in=project_ids.split(','))
            for p in projects:
                rows.append(self._summarize(
                    SiteProgressTask.objects.filter(project=p), p.name,
                    extra={'level': 'Project'}
                ))

        elif scope == 'districts':
            project_id = request.query_params.get('project_id')
            district_ids = request.query_params.get('district_ids')
            districts_qs = SiteProgressTask.objects.filter(project_id=project_id).values_list('district_id', flat=True).distinct()
            districts = District.objects.filter(id__in=districts_qs).order_by('name')
            if district_ids:
                districts = districts.filter(id__in=district_ids.split(','))
            for d in districts:
                rows.append(self._summarize(
                    SiteProgressTask.objects.filter(project_id=project_id, district=d), d.name,
                    extra={'level': 'District'}
                ))

        elif scope == 'sites':
            district_id = request.query_params.get('district_id')
            site_ids = request.query_params.get('site_ids')
            site_id_list = SiteProgressTask.objects.filter(district_id=district_id, site__isnull=False).values_list('site_id', flat=True).distinct()
            sites = Site.objects.filter(id__in=site_id_list).order_by('name')
            if site_ids and site_ids != 'none':
                requested = site_ids.split(',')
                sites = sites.filter(id__in=[s for s in requested if s != 'none'])
            for s in sites:
                rows.append(self._summarize(
                    SiteProgressTask.objects.filter(district_id=district_id, site=s), s.name,
                    extra={'level': 'Site'}
                ))
            include_direct = not site_ids or 'none' in site_ids.split(',')
            if include_direct:
                direct_qs = SiteProgressTask.objects.filter(district_id=district_id, site__isnull=True)
                if direct_qs.exists():
                    rows.append(self._summarize(
                        direct_qs, 'District-wide (no specific site)', extra={'level': 'District-wide'}
                    ))

        return rows, include_breakdown

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        fmt = request.query_params.get('file_format', 'xlsx')
        rows, include_breakdown = self._gather_export_rows(request)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')

        if fmt == 'pdf':
            return self._export_pdf(rows, include_breakdown, timestamp)
        return self._export_xlsx(rows, include_breakdown, timestamp)

    def _export_xlsx(self, rows, include_breakdown, timestamp):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Progress Summary'

        header_fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        headers = ['Level', 'Name', 'Tasks', 'Planned Total', 'Actual Total', '% Complete']
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for row in rows:
            ws.append([
                row.get('level', ''),
                row['name'],
                row['tasks_count'],
                round(row['planned_total'], 2),
                round(row['actual_total'], 2),
                f"{row['percent_complete']}%",
            ])

        if include_breakdown:
            ws.append([])
            ws.append(['Per-Activity Breakdown'])
            ws['A' + str(ws.max_row)].font = Font(bold=True)
            b_headers = ['Scope', 'Activity', 'Unit', 'Planned', 'Actual', '% Complete']
            ws.append(b_headers)
            for cell in ws[ws.max_row]:
                cell.fill = header_fill
                cell.font = header_font
            for row in rows:
                for b in row['breakdown']:
                    ws.append([
                        row['name'], b['label'], b['unit'],
                        round(b['planned'], 2), round(b['actual'], 2), f"{b['percent_complete']}%",
                    ])

        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

        chart_ws = wb.create_sheet('Charts')
        chart_ws.append(['Name', '% Complete'])
        for row in rows:
            chart_ws.append([row['name'], row['percent_complete']])

        bar = BarChart()
        bar.title = 'Overall % Complete'
        bar.y_axis.title = '% Complete'
        bar.x_axis.title = 'Scope'
        data_ref = Reference(chart_ws, min_col=2, min_row=1, max_row=len(rows) + 1)
        cats_ref = Reference(chart_ws, min_col=1, min_row=2, max_row=len(rows) + 1)
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        bar.height = 8
        bar.width = 20
        chart_ws.add_chart(bar, 'D2')

        pie_start_row = len(rows) + 4
        chart_ws.cell(row=pie_start_row, column=1, value='Per-Scope Completion (Actual vs Remaining)')
        pie_row = pie_start_row + 1
        for i, row in enumerate(rows):
            remaining = max(0.0, row['planned_total'] - row['actual_total'])
            label_row = pie_row + i * 4
            chart_ws.cell(row=label_row, column=1, value=row['name'])
            chart_ws.cell(row=label_row + 1, column=1, value='Actual')
            chart_ws.cell(row=label_row + 1, column=2, value=row['actual_total'])
            chart_ws.cell(row=label_row + 2, column=1, value='Remaining')
            chart_ws.cell(row=label_row + 2, column=2, value=remaining)

            pie = PieChart()
            pie.title = row['name']
            data_ref = Reference(chart_ws, min_col=2, min_row=label_row + 1, max_row=label_row + 2)
            cats_ref = Reference(chart_ws, min_col=1, min_row=label_row + 1, max_row=label_row + 2)
            pie.add_data(data_ref)
            pie.set_categories(cats_ref)
            pie.dataLabels = DataLabelList()
            pie.dataLabels.showPercent = True
            pie.height = 6
            pie.width = 8
            chart_ws.add_chart(pie, f'D{label_row}')

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="progress_report_{timestamp}.xlsx"'
        return response

    def _export_pdf(self, rows, include_breakdown, timestamp):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=20 * mm, bottomMargin=20 * mm)
        styles = getSampleStyleSheet()
        elements = [Paragraph('Progress Report', styles['Title']), Spacer(1, 10)]

        summary_data = [['Level', 'Name', 'Tasks', 'Planned', 'Actual', '% Complete']]
        for row in rows:
            summary_data.append([
                row.get('level', ''), row['name'], str(row['tasks_count']),
                f"{round(row['planned_total'], 2):,}", f"{round(row['actual_total'], 2):,}",
                f"{row['percent_complete']}%",
            ])
        summary_table = Table(summary_data, repeatRows=1)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976D2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ]))
        elements.append(summary_table)

        elements.append(Spacer(1, 16))
        elements.append(Paragraph('Overall % Complete by Scope', styles['Heading2']))

        bar_drawing = Drawing(700, 220)
        bar_chart = VerticalBarChart()
        bar_chart.x = 40
        bar_chart.y = 30
        bar_chart.width = 620
        bar_chart.height = 160
        bar_chart.data = [[row['percent_complete'] for row in rows]]
        bar_chart.categoryAxis.categoryNames = [row['name'][:18] for row in rows]
        bar_chart.categoryAxis.labels.angle = 30
        bar_chart.categoryAxis.labels.dy = -10
        bar_chart.valueAxis.valueMin = 0
        bar_chart.valueAxis.valueMax = 100
        bar_chart.bars[0].fillColor = colors.HexColor('#1976D2')
        bar_drawing.add(bar_chart)
        elements.append(bar_drawing)

        elements.append(Spacer(1, 16))
        elements.append(Paragraph('Completion Breakdown by Scope', styles['Heading2']))

        for row in rows:
            remaining = max(0.0, row['planned_total'] - row['actual_total'])
            pie_drawing = Drawing(250, 160)
            pie_drawing.add(String(10, 145, row['name'][:35], fontSize=10, fontName='Helvetica-Bold'))
            pie = Pie()
            pie.x = 60
            pie.y = 10
            pie.width = 110
            pie.height = 110
            pie.data = [row['actual_total'], remaining] if (row['actual_total'] or remaining) else [1]
            pie.labels = ['Actual', 'Remaining'] if (row['actual_total'] or remaining) else ['No data']
            pie.slices[0].fillColor = colors.HexColor('#1976D2')
            if len(pie.data) > 1:
                pie.slices[1].fillColor = colors.HexColor('#E0E0E0')
            pie_drawing.add(pie)
            elements.append(pie_drawing)

        if include_breakdown:
            elements.append(Spacer(1, 16))
            elements.append(Paragraph('Per-Activity Breakdown', styles['Heading2']))
            b_data = [['Scope', 'Activity', 'Unit', 'Planned', 'Actual', '% Complete']]
            for row in rows:
                for b in row['breakdown']:
                    b_data.append([
                        row['name'], b['label'], b['unit'],
                        f"{round(b['planned'], 2):,}", f"{round(b['actual'], 2):,}", f"{b['percent_complete']}%",
                    ])
            b_table = Table(b_data, repeatRows=1)
            b_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#424242')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
            ]))
            elements.append(b_table)

        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="progress_report_{timestamp}.pdf"'
        return response